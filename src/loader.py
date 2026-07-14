"""
CLAUDE L-NACIONAL — Loader v5.1
==============================
Fuente: enloteria.com
  Gana Más:       /resultados-gana-mas
  Nacional Noche: /resultados-nacional-noche

FIXES:
  1. Usa CSS class "numbers" div ID para fecha exacta
     ID formato: lottery_N_numbers_YYYY_MM_DD
  2. La fecha del ID es la fecha REAL del sorteo — se usa TAL CUAL, sin offset.
     (Verificado 1:1 contra enloteria.com para GM y NN, 15 jun–13 jul 2026:
      cero desfases. La instrucción vieja de "restar 1 día" quedó de una
      versión anterior del sitio/scraper y ya NO aplica — se removió del
      docstring en v5.1 porque contradecía el código real y generaba
      confusión. Si enloteria.com vuelve a cambiar su formato de fecha,
      re-verificar contra el sitio antes de reintroducir cualquier offset.)
  3. max_date = ayer para Nacional Noche (el sorteo es a las 9 PM —o 6 PM
     los domingos— y el action corre a las 4 PM, antes del sorteo de hoy).
     max_date = hoy para Gana Más (sorteo 2:30 PM, action corre después).
  4. ensure_updated acepta lottery parameter para actualizar cualquier lotería

NOTA HISTÓRICA (jul 2026): la fila 2026-07-01 Gana Más (86-74-00) fue
borrada en algún momento por parecer "mala". Se confirmó contra el sitio
que SÍ es el resultado real y correcto — probablemente se guardó en su
momento bajo la fecha equivocada (con el offset viejo) y se borró por
error en vez de corregirle la fecha. Ver audit_history.py para detectar
este tipo de fila corrupta automáticamente.
"""

import datetime
import re
import time
from pathlib import Path

import openpyxl
import requests
from bs4 import BeautifulSoup

DATA_DIR     = Path(__file__).parent.parent / "data"
HISTORY_FILE = DATA_DIR / "history.xlsx"
DAYS_ES      = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]

LOTTERY_CANONICAL = {
    "Loteria Nacional- Gana Más":  "Gana Más",
    "Loteria Nacional- Noche":     "Nacional Noche",
    "Gana Más":                    "Gana Más",
    "Nacional Noche":              "Nacional Noche",
}

SCRAPE_URLS = {
    "Gana Más":       "https://enloteria.com/resultados-gana-mas",
    "Nacional Noche": "https://enloteria.com/resultados-nacional-noche",
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "es-DO,es;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

EXCEL_NAMES = {
    "Gana Más":       "Loteria Nacional- Gana Más",
    "Nacional Noche": "Loteria Nacional- Noche",
}


# ── Load Excel ────────────────────────────────────────────────────────────────
def load_history(lottery: str) -> list[dict]:
    wb = openpyxl.load_workbook(HISTORY_FILE, read_only=True, data_only=True)
    ws = wb.active
    draws = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        fecha, sorteo, p1, p2, p3 = row[0], row[1], row[2], row[3], row[4]
        canonical = LOTTERY_CANONICAL.get(str(sorteo).strip())
        if canonical != lottery: continue
        if isinstance(fecha, datetime.datetime):
            date_str = fecha.date().isoformat()
        elif isinstance(fecha, datetime.date):
            date_str = fecha.isoformat()
        else:
            date_str = str(fecha).strip()
        try:
            draws.append({
                "date":     date_str,
                "day_name": DAYS_ES[datetime.date.fromisoformat(date_str).weekday()],
                "lottery":  lottery,
                "p1":       str(int(float(p1))).zfill(2),
                "p2":       str(int(float(p2))).zfill(2),
                "p3":       str(int(float(p3))).zfill(2),
            })
        except (ValueError, TypeError):
            continue
    wb.close()
    draws.sort(key=lambda x: x["date"])

    # Filter holiday duplicates
    filtered = []
    for i, d in enumerate(draws):
        if i == 0:
            filtered.append(d)
            continue
        prev = draws[i - 1]
        if d["p1"]==prev["p1"] and d["p2"]==prev["p2"] and d["p3"]==prev["p3"]:
            print(f"  ⚠️  Feriado omitido: {d['date']} {lottery} ({d['p1']}-{d['p2']}-{d['p3']})")
            continue
        filtered.append(d)
    return filtered


def get_last_date(lottery: str) -> str | None:
    draws = load_history(lottery)
    return draws[-1]["date"] if draws else None


# ── Scraper ───────────────────────────────────────────────────────────────────
def fetch_page(url: str, retries: int = 3) -> BeautifulSoup | None:
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=20)
            resp.raise_for_status()
            return BeautifulSoup(resp.text, "html.parser")
        except Exception as e:
            print(f"    ⚠️  Intento {attempt+1}/{retries}: {e}")
            if attempt < retries - 1:
                time.sleep(3 * (attempt + 1))
    return None


def scrape_recent(lottery: str) -> list[dict]:
    """
    Parse enloteria.com using CSS classes.

    HTML structure confirmed:
      <div class="numbers" id="lottery_0_numbers_2026_07_12">
        <div class="result-number">86</div>
        <div class="result-number">79</div>
        <div class="result-number">25</div>
      </div>

    The date in the div ID is the CORRECT real draw date — used as-is.
    Verified 1:1 against the site's own displayed calendar labels for GM
    and NN across 15 jun–13 jul 2026 with zero mismatches. Do not
    reintroduce a +/-1 day offset without re-verifying against the site.
    """
    url  = SCRAPE_URLS[lottery]
    soup = fetch_page(url)
    if not soup:
        return []

    results = []

    for numbers_div in soup.find_all("div", class_="numbers"):
        div_id = numbers_div.get("id", "")
        # Extract date from ID: lottery_N_numbers_YYYY_MM_DD
        m = re.search(r"numbers_(\d{4})_(\d{2})_(\d{2})$", div_id)
        if not m:
            continue
        try:
            # Date in ID is the CORRECT real draw date — use as-is
            real_date = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
            datetime.date.fromisoformat(real_date)  # validate
        except ValueError:
            continue

        # Extract numbers from result-number divs
        num_divs = numbers_div.find_all("div", class_="result-number")
        nums = []
        for nd in num_divs:
            t = nd.get_text(strip=True)
            if re.match(r"^\d{1,2}$", t) and 0 <= int(t) <= 99:
                nums.append(str(int(t)).zfill(2))

        if len(nums) == 3:
            results.append({
                "date":     real_date,
                "day_name": DAYS_ES[datetime.date.fromisoformat(real_date).weekday()],
                "lottery":  lottery,
                "p1": nums[0], "p2": nums[1], "p3": nums[2],
            })

    # Deduplicate by date
    seen = {}
    for r in results:
        if r["date"] not in seen:
            seen[r["date"]] = r
    return sorted(seen.values(), key=lambda x: x["date"])


def append_to_excel(new_rows: list[dict]) -> int:
    if not new_rows:
        return 0
    wb = openpyxl.load_workbook(HISTORY_FILE)
    ws = wb.active

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]: continue
        if isinstance(row[0], (datetime.datetime, datetime.date)):
            fecha = row[0].date().isoformat() if isinstance(row[0], datetime.datetime) else row[0].isoformat()
        else:
            fecha = str(row[0]).strip()
        canonical = LOTTERY_CANONICAL.get(str(row[1]).strip(), str(row[1]).strip())
        existing.add((fecha, canonical))

    added = 0
    for r in sorted(new_rows, key=lambda x: x["date"]):
        key = (r["date"], r["lottery"])
        if key in existing:
            continue
        ws.append([r["date"], EXCEL_NAMES[r["lottery"]], r["p1"], r["p2"], r["p3"]])
        existing.add(key)
        added += 1

    if added > 0:
        wb.save(HISTORY_FILE)
        print(f"  ✅ {added} nuevos resultados guardados en history.xlsx")
    else:
        print(f"  ℹ️  Sin cambios en history.xlsx")
    wb.close()
    return added


def ensure_updated(lottery: str) -> list[dict]:
    """
    Update the history for a given lottery.

    Gana Más: draws 2:30 PM, workflow runs 4 PM RD -> today's result is
    already out, so max_date = today.

    Nacional Noche: draws 9 PM (6 PM Sundays), workflow runs 4 PM RD ->
    today's result is NOT out yet, so max_date = yesterday.
    """
    last_date = get_last_date(lottery)
    today     = datetime.date.today().isoformat()
    yesterday = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()

    max_date = today if lottery == "Gana Más" else yesterday

    if last_date == max_date:
        print(f"  ✅ {lottery}: historial al día ({last_date})")
        return load_history(lottery)

    print(f"  📥 {lottery}: último={last_date}, scrapeando hasta {max_date}...")
    scraped = scrape_recent(lottery)

    new_rows = [
        r for r in scraped
        if r["date"] > (last_date or "2000-01-01")
        and r["date"] <= max_date
    ]

    if new_rows:
        for r in new_rows:
            print(f"    ✅ {r['date']} [{lottery}]: {r['p1']}-{r['p2']}-{r['p3']}")
        append_to_excel(new_rows)
    else:
        print(f"  ⚠️  No se encontraron resultados nuevos para {lottery}")

    return load_history(lottery)